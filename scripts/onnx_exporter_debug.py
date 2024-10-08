import onnx
import openpyxl
from onnx import shape_inference

def export_onnx_to_excel(onnx_path, output_excel_path):
    # 载入ONNX模型并推断形状
    model = onnx.load(onnx_path)
    inferred_model = shape_inference.infer_shapes(model)
    graph = inferred_model.graph

    # 初始化 Excel 工作簿
    wb = openpyxl.Workbook()

    # Sheet1: 去重后的算子类型
    ws1 = wb.active
    ws1.title = "Unique Operators"

    # Sheet2: 每一层的详细信息
    ws2 = wb.create_sheet("Layer Details")

    # Sheet3: 模型的输入和输出信息
    ws3 = wb.create_sheet("Model IO Shapes")

    # 初始化存储信息的集合
    unique_operators = set()
    layer_details = []

    # 构建一个字典来存储每个值的形状
    value_shapes = {}
    for value_info in graph.value_info:
        shape = [dim.dim_value if dim.dim_value > 0 else -1 for dim in value_info.type.tensor_type.shape.dim]
        value_shapes[value_info.name] = shape

    # 获取输入和输出的形状
    def get_shapes(names):
        shapes = []
        for name in names:
            if name in value_shapes:
                shapes.append(value_shapes[name])
            else:
                shapes.append([-1])
        return shapes

    # 获取模型的输入和输出信息
    def get_model_io_shapes(io_list):
        io_shapes = []
        for io in io_list:
            shape = [dim.dim_value if dim.dim_value > 0 else -1 for dim in io.type.tensor_type.shape.dim]
            io_shapes.append((io.name, shape))
        return io_shapes

    # 模型的输入信息
    input_shapes = get_model_io_shapes(graph.input)
    output_shapes = get_model_io_shapes(graph.output)

    # 遍历 ONNX 模型中的节点
    for node in graph.node:
        operator_type = node.op_type
        inputs = node.input
        outputs = node.output

        # 获取输入和输出的形状
        input_shapes = get_shapes(inputs)
        output_shapes = get_shapes(outputs)

        # 将算子类型加入到去重集合中
        unique_operators.add(operator_type)

        # 将详细信息存储在列表中
        layer_details.append([
            operator_type,
            ",".join(inputs),
            str(input_shapes),
            ",".join(outputs),
            str(output_shapes)
        ])

    # 将去重后的算子类型写入 Sheet1
    ws1.append(["Unique Operators"])
    for operator in unique_operators:
        ws1.append([operator])

    # 将每一层的详细信息写入 Sheet2
    ws2.append(["Operator Type", "Inputs", "Input Shapes", "Outputs", "Output Shapes"])
    for detail in layer_details:
        ws2.append(detail)

    # 将模型的输入和输出信息写入 Sheet3
    ws3.append(["Model Input Name", "Input Shape"])
    print(input_shapes)
    for name, shape in input_shapes:
        shape_str = ' '.join(map(str, shape)) if isinstance(shape, list) else str(shape)
        ws3.append([name, shape_str])

    ws3.append([])  # 空行分隔
    ws3.append(["Model Output Name", "Output Shape"])
    for name, shape in output_shapes:
        shape_str = ' '.join(map(str, shape)) if isinstance(shape, list) else str(shape)
        ws3.append([name, shape_str])

    # 保存Excel文件
    wb.save(output_excel_path)
    print(f"Excel file saved to {output_excel_path}")

# 使用该函数
if __name__ == '__main__':
    export_onnx_to_excel('../resource/resnet18-v1-7.onnx', 'onnx_layers_5.xlsx')
